#!/usr/bin/env python3
"""
rebuild_db.py - 从 BIS 官方 Excel 重建内置数据库
用法: python3 rebuild_db.py <Excel文件路径>
"""
import json
import os
import sys

def parse_excel_and_build_db(excel_path):
    """解析 BIS Excel 并生成数据库"""
    import openpyxl
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['中国实体主表']
    
    # Get headers
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 4 and row[0]:
            headers = {j: v for j, v in enumerate(row) if v}
            break
    
    if not headers:
        print("❌ 无法找到表头行")
        return None, None
    
    # Parse records
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 4:
            continue
        rec = {}
        for j, v in enumerate(row):
            if v and headers.get(j):
                rec[headers[j]] = str(v)
        if rec:
            records.append(rec)
    
    print(f"📊 解析到 {len(records)} 条实体记录")
    
    # Build DB
    db = {}
    for rec in records:
        en_name = rec.get('英文名', '')
        cn_name = rec.get('中文名（官方/常用译名；未核实留空）', '')
        date = rec.get('列入时间', '')
        policy = rec.get('许可审查政策（中文概述）', '')
        footnote = rec.get('脚注编号', '')
        reason = rec.get('列入原因（按EAR条款归纳）', '')
        alt_names = rec.get('别名 / Alt Names', '')
        
        entry = {
            'status': '已列入',
            'en_name': en_name,
            'date': date,
            'policy': policy,
            'footnote': footnote if footnote else '-',
            'reason': reason,
            'source': 'BIS官方Excel'
        }
        
        if en_name:
            db[en_name] = entry
        if cn_name:
            db[cn_name] = entry
        if alt_names:
            for alt in alt_names.split(';'):
                alt = alt.strip()
                if alt:
                    db[alt] = entry
    
    print(f"📊 生成 {len(db)} 条映射（含别名）")
    return db, records


def load_existing_alias_map():
    """加载现有中文别名映射"""
    alias_file = os.path.join(os.path.dirname(__file__), 'cache', 'chinese_alias_map.json')
    if os.path.exists(alias_file):
        with open(alias_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def main():
    if len(sys.argv) < 2:
        print("用法: python3 rebuild_db.py <Excel文件路径>")
        print("示例: python3 rebuild_db.py ~/Downloads/BIS_Entity_List_China_latest.xlsx")
        return
    
    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        return
    
    print(f"📂 读取 Excel: {excel_path}")
    
    db, records = parse_excel_and_build_db(excel_path)
    if not db:
        return
    
    # Load existing alias map to preserve manually added entries
    existing_alias = load_existing_alias_map()
    print(f"📊 保留现有别名映射: {len(existing_alias)} 条")
    
    # Merge: use existing alias map values to add to db
    # (existing aliases that map to None = explicitly not listed)
    
    # Save DB to skill's cache directory
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cache_dir = os.path.join(skill_root, 'cache')
    os.makedirs(cache_dir, exist_ok=True)
    
    db_file = os.path.join(cache_dir, 'bis_china_official_db.json')
    with open(db_file, 'w', encoding='utf-8') as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
    print(f"✅ 已保存数据库: {db_file}")
    print(f"   共 {len(db)} 条映射，{len(records)} 家实体")
    
    # Save alias map
    alias_file = os.path.join(cache_dir, 'chinese_alias_map.json')
    with open(alias_file, 'w', encoding='utf-8') as f:
        json.dump(existing_alias, f, ensure_ascii=False, indent=2)
    print(f"✅ 已保存别名映射: {alias_file}")


if __name__ == '__main__':
    main()
